#!/usr/bin/env python3
"""
dunsteel_variation.py - Create a new variation folder + Excel from the PM08
template on the Dunsteel S: drive, auto-numbered, with the header pre-filled.

Worker for the /dunsteel-variation Claude Code skill.

Behaviour:
  1. Find the project folder on S: drive by number prefix ("[NNN] -*").
  2. Navigate to "[project]\\04 Variations\\".
  3. Scan existing "V[NN] ..." folders, find the highest number, increment.
  4. Create a new folder "V[NN] [title]\\".
  5. Copy the PM08 template into it, renamed "V[NN] [title].xlsx".
  6. Pre-fill only the named header cells via openpyxl (never formulas/tables).
  7. Open the file via the Windows `start` command.

Cell mapping (verified against the live PM08 template AND existing 501
V04 / V08 / V09 files - the real files are the source of truth):
  A3  CLIENT          (merged A3:C3,  referenced downstream by D65 =A3)
  A5  Job Name / code (merged A5:C5,  template default placeholder "JOBCODE")
  D5  Job Address     (merged D5:F5)
  B11 Job No.         (merged B11:C11, integer)
  B12 Variation #     (merged B12:C12, integer)
  C13 Variation Title (merged C13:I13)
  B8  Date In         already "=TODAY()" - left untouched
  B9  Revision        already "A" - left untouched

Hard rule: no em dashes anywhere in this file or in any generated content.

Usage:
  python scripts/dunsteel_variation.py <project_number> <variation title...>
  python scripts/dunsteel_variation.py 501 HV Room Stair Access Modifications

Flags:
  --no-open                   do not auto-open the file after creation
  --variations-dir "PATH"     override the "04 Variations" directory
                              (used by the test harness to target an isolated
                              TEST_ subfolder; never used in normal operation)
  --json                      emit a machine-readable JSON result block

Exit codes:
  0  success
  2  project not found / ambiguous
  3  template not found
  4  bad arguments
  5  other failure
"""

import argparse
import json
import re
import subprocess
import sys
import warnings
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.stderr.write(
        "openpyxl is not installed. Run: pip install openpyxl\n"
    )
    sys.exit(5)

# openpyxl warns about the data-validation extension in this template; harmless.
warnings.simplefilter("ignore")

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

CURRENT_PROJECTS_ROOT = Path(r"S:\Operations\01 Current Project")
PM08_TEMPLATE = Path(
    r"S:\Operations\11 Project Management Database"
    r"\Project Management Templates\PM08. Variation Template.xlsx"
)
VARIATIONS_SUBFOLDER = "04 Variations"

# Per-project header facts. Keyed by project number (string).
# Job No. and Variation # are written as integers to match the existing files.
PROJECT_HEADERS = {
    "501": {
        "client": "Northbridge Constructions",
        "job_name": "ATSYD02",
        "job_address": "[site address]",
    },
}

# Characters Windows does not allow in file/folder names.
_INVALID_WIN_CHARS = r'<>:"/\|?*'
_INVALID_WIN_RE = re.compile(r'[<>:"/\\|?*\x00-\x1f]')


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def sanitise_title(title: str):
    """Return (clean_title, warnings list) safe for a Windows file/folder name.

    Strips characters Windows forbids, collapses whitespace, and trims any
    trailing dots or spaces (also illegal at the end of a Windows name).
    """
    warns = []
    bad = sorted({c for c in title if _INVALID_WIN_RE.match(c)})
    if bad:
        warns.append(
            "Removed characters invalid in Windows filenames: "
            + " ".join(repr(c) for c in bad)
        )
    clean = _INVALID_WIN_RE.sub("", title)
    clean = re.sub(r"\s+", " ", clean).strip()
    clean = clean.rstrip(". ")
    if not clean:
        warns.append("Title was empty after sanitising.")
    elif clean != title.strip():
        warns.append(f"Title normalised to: {clean!r}")
    return clean, warns


def find_project_folder(number: str):
    """Find the single project folder matching '[number] -*'. Returns Path.

    Raises LookupError if zero or more than one match (caller reports clean).
    """
    if not CURRENT_PROJECTS_ROOT.exists():
        raise LookupError(
            f"Current Project root not accessible: {CURRENT_PROJECTS_ROOT}"
        )
    matches = [
        p for p in CURRENT_PROJECTS_ROOT.iterdir()
        if p.is_dir() and re.match(rf"^{re.escape(number)}\s*-", p.name)
    ]
    if not matches:
        raise LookupError(
            f"No project folder found for number {number!r} under "
            f"{CURRENT_PROJECTS_ROOT}"
        )
    if len(matches) > 1:
        names = ", ".join(repr(m.name) for m in matches)
        raise LookupError(
            f"Multiple project folders match {number!r}: {names}. "
            "Refusing to guess."
        )
    return matches[0]


def next_variation_number(variations_dir: Path):
    """Scan 'V[NN] ...' folders, return next integer. VX folders are ignored."""
    highest = 0
    pat = re.compile(r"^V(\d+)\b", re.IGNORECASE)
    for entry in variations_dir.iterdir():
        if not entry.is_dir():
            continue
        m = pat.match(entry.name)
        if m:
            highest = max(highest, int(m.group(1)))
    return highest + 1


def find_folder_by_title(variations_dir: Path, clean_title: str):
    """Return an existing 'V[NN] <title>' folder matching this exact title.

    Used for idempotency: rerunning with the same title must reuse the folder
    already created for it rather than allocating a fresh number. Comparison is
    case-insensitive on the title portion after the 'V[NN] ' prefix.
    """
    pat = re.compile(r"^V(\d+)\s+(.*)$", re.IGNORECASE)
    target = clean_title.casefold()
    for entry in variations_dir.iterdir():
        if not entry.is_dir():
            continue
        m = pat.match(entry.name)
        if m and m.group(2).casefold() == target:
            return entry, int(m.group(1))
    return None, None


def prefill_header(xlsx_path: Path, headers: dict, var_number: int,
                   var_title: str, job_no: str):
    """Write only the named header cells. Never touches formulas or tables."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    ws = wb["Variation"] if "Variation" in wb.sheetnames else wb.active

    ws["A3"] = headers["client"]            # CLIENT  (D65 mirrors this)
    ws["A5"] = headers["job_name"]          # Job Name / code
    ws["D5"] = headers["job_address"]       # Job Address
    ws["B11"] = int(job_no) if str(job_no).isdigit() else job_no  # Job No.
    ws["B12"] = var_number                  # Variation #
    ws["C13"] = var_title                   # Variation Title
    # B8 (Date In =TODAY()) and B9 (Revision "A") are template defaults: leave.

    wb.save(xlsx_path)


def open_file(path: Path):
    """Open the file with its default application via Windows `start`."""
    # `start` is a cmd builtin; empty "" is the required window-title arg.
    subprocess.run(["cmd", "/c", "start", "", str(path)], check=False)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main(argv=None):
    parser = argparse.ArgumentParser(
        description="Create a new Dunsteel variation folder + PM08 Excel."
    )
    parser.add_argument("number", help="Project number, e.g. 501")
    parser.add_argument("title", nargs="+", help="Variation title")
    parser.add_argument("--no-open", action="store_true",
                        help="Do not auto-open the created file")
    parser.add_argument("--variations-dir", default=None,
                        help="Override the 04 Variations directory (testing)")
    parser.add_argument("--json", action="store_true",
                        help="Emit a JSON result block")
    args = parser.parse_args(argv)

    result = {"ok": False, "warnings": []}

    def finish(code, **extra):
        result.update(extra)
        if args.json:
            print("RESULT_JSON " + json.dumps(result))
        return code

    number = args.number.strip()
    raw_title = " ".join(args.title).strip()

    clean_title, title_warns = sanitise_title(raw_title)
    result["warnings"].extend(title_warns)
    if not clean_title:
        sys.stderr.write("Variation title is empty after sanitising.\n")
        return finish(4, error="empty_title")

    headers = PROJECT_HEADERS.get(number)
    if headers is None:
        # Safe fallback: still create the file, leave header fields blank so
        # Nathan fills them. Report it loudly rather than guessing wrong data.
        result["warnings"].append(
            f"No stored header facts for project {number}. CLIENT, Job Name "
            "and Job Address left blank for manual entry."
        )
        headers = {"client": "", "job_name": "", "job_address": ""}

    # Resolve the variations directory.
    if args.variations_dir:
        variations_dir = Path(args.variations_dir)
        project_folder = variations_dir.parent
        if not variations_dir.exists():
            sys.stderr.write(f"Override variations dir not found: "
                             f"{variations_dir}\n")
            return finish(2, error="variations_dir_missing")
    else:
        try:
            project_folder = find_project_folder(number)
        except LookupError as e:
            sys.stderr.write(str(e) + "\n")
            return finish(2, error="project_not_found", detail=str(e))
        variations_dir = project_folder / VARIATIONS_SUBFOLDER
        if not variations_dir.exists():
            sys.stderr.write(
                f"'{VARIATIONS_SUBFOLDER}' folder not found in "
                f"{project_folder}\n"
            )
            return finish(2, error="no_variations_folder")

    result["project_folder"] = str(project_folder)
    result["variations_dir"] = str(variations_dir)

    if not PM08_TEMPLATE.exists():
        sys.stderr.write(f"PM08 template not found: {PM08_TEMPLATE}\n")
        return finish(3, error="template_not_found")

    # Idempotency first: if a folder for THIS exact title already exists at any
    # V number, reuse it rather than allocating a new number.
    existing_folder, existing_num = find_folder_by_title(
        variations_dir, clean_title
    )
    if existing_folder is not None:
        var_number = existing_num
        vlabel = f"V{var_number:02d}"
        new_folder = existing_folder
        folder_name = existing_folder.name
        result["warnings"].append(
            f"A folder for this title already exists ({existing_folder.name!r}); "
            "reusing it instead of creating a duplicate."
        )
    else:
        var_number = next_variation_number(variations_dir)
        vlabel = f"V{var_number:02d}"
        folder_name = f"{vlabel} {clean_title}"
        new_folder = variations_dir / folder_name
        new_folder.mkdir(parents=True, exist_ok=False)

    new_file = new_folder / f"{folder_name}.xlsx"

    result["folder"] = str(new_folder)
    result["file"] = str(new_file)
    result["variation_label"] = vlabel
    result["variation_number"] = var_number

    if new_file.exists():
        result["warnings"].append(
            "Variation file already exists; left untouched (idempotent)."
        )
    else:
        # Copy the template in, then pre-fill. Copy via openpyxl save would
        # drop the macro/validation extensions, so do a raw byte copy first.
        new_file.write_bytes(PM08_TEMPLATE.read_bytes())
        prefill_header(new_file, headers, var_number, clean_title, number)

    result["ok"] = True

    if not args.no_open and new_file.exists():
        try:
            open_file(new_file)
            result["opened"] = True
        except Exception as e:  # opening is best-effort, never fatal
            result["warnings"].append(f"Could not auto-open file: {e}")
            result["opened"] = False

    # Human-readable summary.
    print(f"Variation created: {vlabel} {clean_title}")
    print(f"  Project folder : {project_folder}")
    print(f"  Folder         : {new_folder}")
    print(f"  File           : {new_file}")
    if result["warnings"]:
        print("  Warnings:")
        for w in result["warnings"]:
            print(f"    - {w}")

    return finish(0)


if __name__ == "__main__":
    sys.exit(main())
